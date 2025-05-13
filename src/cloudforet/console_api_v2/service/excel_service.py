import logging
import json
import uuid
import pytz
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from typing import Optional, Any, Union, List, Dict
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from starlette.responses import StreamingResponse

from spaceone.core import cache
from spaceone.core import utils
from spaceone.core.service import *
from spaceone.core.error import ERROR_CACHE_CONFIGURATION

from cloudforet.console_api_v2.manager.cloudforet_manager import CloudforetManager
from cloudforet.console_api_v2.model.excel.request import ExcelExportRequest

_LOGGER = logging.getLogger(__name__)


@authentication_handler
@event_handler
class ExcelService(BaseService):
    resource = "Excel"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    @transaction()
    def export(self, params: dict) -> dict:
        if cache.is_set(alias="local"):
            self._check_excel_export_params(params)

            redis_key = uuid.uuid4()
            key_dict = {
                "request_body": params,
                "auth_info": {
                    "token": self.metadata.get("token"),
                    "domain_id": self.transaction.get_meta("authorization.domain_id"),
                    "audience": self.transaction.get_meta("authorization.audience"),
                    "owner_type": self.transaction.get_meta("authorization.owner_type"),
                },
            }

            redis_value = json.dumps(key_dict)
            cache.set(
                key=f"console-api:excel:{redis_key}", value=redis_value, alias="local"
            )
        else:
            raise ERROR_CACHE_CONFIGURATION()

        return {"file_link": f"/console-api/extension/excel/download?key={redis_key}"}

    @transaction()
    @convert_model
    def download(self, params: ExcelExportRequest) -> StreamingResponse:

        params = params.dict()
        wb = self._create_excel(params)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        file_name = self._get_file_name(params)

        headers = {
            "Content-Disposition": f"attachment; filename={file_name}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @staticmethod
    def _check_excel_export_params(options: dict, array_idx: int = None):
        source = options.get("source")
        template = options.get("template")

        def idx_msg():
            return f" in {array_idx}th index" if array_idx is not None else ""

        if not source:
            raise Exception(f"Required Parameter. (key = source{idx_msg()})")
        if not template:
            raise Exception(f"Required Parameter. (key = template{idx_msg()})")

        if "data" not in source:
            if not source.get("url") or not source.get("param"):
                raise Exception(
                    f"Invalid Parameter. (source{idx_msg()} = must have url and param keys)"
                )
            if not isinstance(source["url"], str):
                raise Exception(
                    f"Parameter type is invalid. (source.url{idx_msg()} = string)"
                )
            if "query" not in source["param"]:
                raise Exception(
                    f"Invalid Parameter. (source.param{idx_msg()} = must have query)"
                )

    def _create_excel(self, excel_options: Union[dict, List[dict]]) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        if isinstance(excel_options, list):
            for opt in excel_options:
                self._create_worksheet(wb, opt)
        else:
            self._create_worksheet(wb, excel_options)

        return wb

    def _create_worksheet(self, workbook: Workbook, excel_option: dict) -> Worksheet:
        template = excel_option.get("template", {})
        options = template.get("options", {})
        sheet_name = options.get("sheet_name", "Sheet1")
        fields = template.get("fields", [])
        header_message = options.get("header_message", {})

        source = excel_option.get("source", {})

        # New worksheet
        ws = workbook.create_sheet(title=sheet_name)

        # set worksheet headers
        columns = self._get_excel_columns(fields)
        self._set_excel_header(ws, columns, header_message)

        # add rows at worksheet
        if source:
            ws = self._set_excel_cell_rows(ws, source, fields, options)

        # set worksheet style
        self._set_worksheet_row_style(ws, header_message)
        self._set_worksheet_column_style(ws, header_message)

        return ws

    def _set_excel_cell_rows(
        self, worksheet: Worksheet, source: dict, fields: list, options: dict
    ) -> Worksheet:
        cf_mgr = CloudforetManager()
        token = self.transaction.get_meta("authorization.token")
        reference_resource_map = self._get_reference_resource_map(cf_mgr, fields, token)

        timezone = options.get("timezone", "UTC")

        if data := source.get("data"):
            raw_data = data
        else:
            raw_data = self._get_raw_data(cf_mgr, source, token)

        # add rows
        for row in raw_data:
            excel_row_data = {}

            for field in fields:
                key = field["key"]
                name = field.get("name") or key
                field_type = field.get("type")
                reference = field.get("reference")
                options = field.get("options", {})
                enum_items = field.get("enum_items", {})

                row_value = self.get_value_by_path(row, key)

                if reference:
                    resource_type = reference["resource_type"]
                    reference_key = reference["reference_key"]
                    reference_resources = reference_resource_map.get(resource_type, [])
                    row_value = self._convert_value_with_reference_data(
                        row_value, reference_key, reference_resources
                    )

                if row_value is None:
                    value = ""
                elif field_type == "datetime":
                    dt = datetime.fromisoformat(row_value)
                    tz = pytz.timezone(timezone)
                    value = dt.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S")

                elif field_type == "currency":
                    value = row_value
                elif field_type == "enum":
                    value = enum_items.get(row_value, row_value)
                elif field_type == "list":
                    value = ", ".join(row_value)

                elif isinstance(row_value, list):
                    value = "\n".join([str(v) for v in row_value])
                elif isinstance(row_value, int) or isinstance(row_value, float):
                    value = row_value
                else:
                    value = str(row_value)

                excel_row_key = field.get("name", key)
                excel_row_data[excel_row_key] = value

            excel_data_values = []
            for value in excel_row_data.values():
                excel_data_values.append(value)

            worksheet.append(excel_data_values)
        return worksheet

    def _set_excel_header(
        self, worksheet: Worksheet, columns: list, header_message: dict
    ):
        header_title = header_message.get("title")
        header_row_number = 2 if header_title else 1

        headers = [column["header"] for column in columns]
        if header_title:
            self._set_excel_header_message(worksheet, header_title)
            for col_index, header in enumerate(headers, start=1):
                worksheet.cell(row=header_row_number, column=col_index, value=header)
        else:
            worksheet.append(headers)

        # set header style
        self._set_worksheet_header_style(worksheet, header_row_number, len(columns))

        # set column width
        ExcelService._set_worksheet_column_style(worksheet, header_message)

    def _set_excel_header_message(self, worksheet: Worksheet, title: str):
        worksheet.insert_rows(1)
        worksheet["A1"].value = title
        self._set_header_message_style(worksheet)

    @staticmethod
    def _set_header_message_style(worksheet: Worksheet):
        cell = worksheet["A1"]
        cell.font = Font(bold=True, size=22, color="003566")
        cell.alignment = Alignment(vertical="bottom", horizontal="left")

    @staticmethod
    def _get_excel_columns(fields: List[dict]) -> List[Dict]:
        return [
            {
                "header": field["name"],
                "key": field["key"],
                "height": 24,
                "style": {
                    "font": {"size": 12},
                    "alignment": {
                        "vertical": "top",
                        "horizontal": "left",
                        "wrapText": True,
                    },
                },
            }
            for field in fields
        ]

    def _set_worksheet_header_style(
        self, worksheet: Worksheet, header_row_number: int, column_length: int
    ):
        for num in range(0, column_length):
            column_letter = self._get_column_letter(num)
            cell = worksheet[f"{column_letter}{header_row_number}"]

            cell.fill = PatternFill(fill_type="solid", fgColor="003566")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left")

    @staticmethod
    def _get_column_letter(number: int) -> str:
        letters = ""
        while number >= 0:
            letters = chr(number % 26 + ord("A")) + letters
            number = number // 26 - 1
        return letters

    @staticmethod
    def _set_worksheet_column_style(worksheet: Worksheet, header_message: dict = None):
        header_row_number = 2 if header_message and header_message.get("title") else 1
        min_width = 10

        for column_cells in worksheet.columns:
            max_column_length = 0
            for cell in column_cells:
                if cell.row >= header_row_number:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_column_length = max(
                        max_column_length, min_width, len(cell_value)
                    )
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = max_column_length + 2

    def _get_reference_resource_map(
        self, cloudforet_mgr: CloudforetManager, fields: list, token: str
    ) -> dict:
        reference_resource_map = {}

        references_info = []
        for field in fields:
            reference = field.get("reference", {}) or {}
            resource_type = reference.get("resource_type")

            if (
                reference
                and resource_type
                and resource_type not in reference_resource_map
            ):
                references_info.append(reference)

        # get reference resources
        for reference_info in references_info:
            resource_type = reference_info.get("resource_type")
            if resource_type in reference_resource_map:
                continue
            try:
                res = self._get_reference_resources(
                    cloudforet_mgr, reference_info, token
                )
                reference_resource_map[resource_type] = res
            except Exception as e:
                _LOGGER.error(f"Failed to get reference resources: {e}", exc_info=True)
        return reference_resource_map

    @staticmethod
    def _get_reference_resources(
        cloudforet_mgr: CloudforetManager, reference_info: dict, token: str
    ) -> list:
        grpc_method = f"{reference_info.get('resource_type')}.list"
        reference_key = reference_info.get("reference_key")
        query_params = {"query": {"only": [reference_key, "name"]}}
        results = []

        responses = list(cloudforet_mgr.paginate_api(grpc_method, query_params, token))
        for res in responses:
            results.extend(res.get("results", []))

        return results

    @staticmethod
    def _get_raw_data(
        cloudforet_mgr: CloudforetManager, source: dict, token: str
    ) -> list:

        url = source.get("url")
        params = source.get("param", {})
        results = []

        grpc_method: str = cloudforet_mgr.convert_grpc_method_from_url(url)

        responses = list(cloudforet_mgr.paginate_api(grpc_method, params, token))
        for res in responses:
            results.extend(res.get("results", []))

        return results

    @staticmethod
    def _convert_value_with_reference_data(
        row_value: Any, reference_key: str, reference_resources: list
    ) -> Union[str, list]:

        if isinstance(row_value, list):
            for idx, value in enumerate(row_value):
                for reference_resource in reference_resources:
                    if reference_resource.get(reference_key) == value:
                        row_value[idx] = reference_resource.get("name")
                        break
        else:
            for reference_resource in reference_resources:
                if reference_resource.get(reference_key) == row_value:
                    row_value = reference_resource.get("name")
                    break

        return row_value

    @staticmethod
    def _get_file_name(excel_options: Union[dict, List[dict]]) -> str:
        if isinstance(excel_options, list):
            options = excel_options[0]
        else:
            options = excel_options

        timezone = options.get("template", {}).get("options", {}).get("timezone", "UTC")
        prefix = (
            options.get("template", {})
            .get("options", {})
            .get("file_name_prefix", "export")
        )

        now = datetime.now(pytz.timezone(timezone)).strftime("%Y%m%d")
        return f"{prefix}_export_{now}.xlsx"

    @staticmethod
    def _get_row_value(row_value: Union[str, list, dict], key: str) -> Any:
        if isinstance(row_value, dict):
            print("row_value is dict")
            row_value = utils.get_dict_value(data=row_value, dotted_key=key)
        elif isinstance(row_value, list):
            row_value = utils.get_list_values(row_value, key)

        return row_value

    def get_value_by_path(
        self, data: Any, path: Optional[str], depth: Optional[int] = None
    ) -> Any:
        if not isinstance(path, str):
            return data

        target = data
        path_arr = path.split(".")

        last_depth_key = None
        if depth is not None:
            last_depth_key = ".".join(path_arr[depth:])
            path_arr = path_arr[:depth]

        path_count = depth if depth is not None else len(path_arr)

        for i in range(path_count):
            if target is None:
                return None

            current_path = path_arr[i]

            # 숫자인 경우 list index로 처리
            if isinstance(target, list):
                try:
                    index = int(current_path)
                    target = target[index]
                except (ValueError, IndexError):
                    return None
            elif isinstance(target, dict):
                target = target.get(current_path)
            else:
                return None

        if depth and last_depth_key:
            return self.get_value_by_path(target, last_depth_key)

        return target

    @staticmethod
    def _set_worksheet_row_style(worksheet: Worksheet, header_message: dict = None):
        header_row_number = 2 if header_message and header_message.get("title") else 1

        thin_border = Border(
            top=Side(style="thin", color="E5E5E8"),
            left=Side(style="thin", color="E5E5E8"),
            bottom=Side(style="thin", color="E5E5E8"),
            right=Side(style="thin", color="E5E5E8"),
        )

        striped_fill = PatternFill(fill_type="solid", fgColor="F7F7F7")

        for row in worksheet.iter_rows():
            row_number = row[0].row
            for cell in row:
                cell.border = thin_border
            if row_number > header_row_number and row_number % 2 == 0:
                for cell in row:
                    cell.fill = striped_fill
